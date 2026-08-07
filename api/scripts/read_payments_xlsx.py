"""Читает листы поступлений из PaymentsBattery и складывает их в JSON для загрузчика.

Какие листы и что с них берём — решение Николая 07.08.2026:

    РС-Поступ-Хольц   основная и финансовая деятельность
    РС-Поступ-ИП      только основная деятельность
    ПоступленияН      только основная деятельность

Строки других статей не выбрасываются, а помечаются признаком «исключена»:
загрузчик занесёт их в промежуточную таблицу, но в платежи не переведёт.
Так в базе остаётся история — видно, что было в файле и почему не взято, — и
при этом снятие своих денег в кассу не попадает в выручку. Выбрасывать их
совсем нельзя: тогда через месяц никто не вспомнит, была строка или нет.

Колонки берём по заголовкам, а не по буквам: листов три, и вставка столбца в
любом из них иначе тихо сдвинула бы разбор.

Ничего не пишет в базу. На выходе — файл JSON.

Использование:
    python read_payments_xlsx.py PaymentsBattery.xlsx payments.json
"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# лист → статьи поступлений, которые переводим в платежи
SHEETS = {
    "РС-Поступ-Хольц": {"основная деятельность", "финансовая деятельность"},
    "РС-Поступ-ИП": {"основная деятельность"},
    "ПоступленияН": {"основная деятельность"},
}

COLUMNS = {
    "№": "sequence",
    "плательщик": "counterparty",
    "сумма": "amount",
    "пояснение": "explanation",
    "основание оплаты": "basis",
    "клиент": "client",
    "дата оплаты": "date",
    "статья поступлений": "category",
    "период": "period",
    "план": "plan",
    "цу": "account",
    "вид движений": "movement",
    "менеджер": "manager",
    "вид выручки": "revenueType",
}


def as_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def as_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def as_amount(value: object) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if value is None:
        return None
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def find_header(sheet) -> tuple[int, dict[int, str]]:
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        titles = {
            str(value).strip().lower(): position
            for position, value in enumerate(row, start=1)
            if value is not None
        }
        if "плательщик" in titles and "сумма" in titles:
            return index, {
                position: COLUMNS[title]
                for title, position in titles.items()
                if title in COLUMNS
            }
    raise SystemExit("не нашлась строка заголовков с колонками «Плательщик» и «Сумма»")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--period-from", default="2026-01-01")
    arguments = parser.parse_args()

    since = date.fromisoformat(arguments.period_from)
    book = load_workbook(arguments.workbook, read_only=True, data_only=True)

    items: list[dict] = []
    for sheet_name, allowed in SHEETS.items():
        if sheet_name not in book.sheetnames:
            print(f"листа «{sheet_name}» нет в книге — пропущен")
            continue
        sheet = book[sheet_name]
        header_row, mapping = find_header(sheet)

        taken = excluded = 0
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if index <= header_row:
                continue
            record = {
                field: (row[position - 1] if position <= len(row) else None)
                for position, field in mapping.items()
            }
            operation_date = as_date(record.get("date"))
            if operation_date is None or date.fromisoformat(operation_date) < since:
                continue
            amount = as_amount(record.get("amount"))
            if amount is None or amount <= 0:
                continue

            category = as_text(record.get("category"))
            is_allowed = (category or "").strip().lower() in allowed
            # «ВП» в статье — внутренний перевод: снятие своих денег в кассу
            # или между своими счетами, выручкой не является
            internal = (category or "").strip().lower().startswith("вп")

            items.append(
                {
                    "sheet": sheet_name,
                    "excelRow": index,
                    "direction": "inflow",
                    "sequence": as_text(record.get("sequence")),
                    "account": as_text(record.get("account")),
                    "counterparty": as_text(record.get("counterparty")),
                    "amount": amount,
                    "date": operation_date,
                    "rawDate": as_text(record.get("date")),
                    "explanation": as_text(record.get("explanation")),
                    "basis": as_text(record.get("basis")),
                    "client": as_text(record.get("client")),
                    "category": category,
                    "period": as_date(record.get("period")),
                    "plan": as_text(record.get("plan")),
                    "movement": as_text(record.get("movement")),
                    "manager": as_text(record.get("manager")),
                    "revenueType": as_text(record.get("revenueType")),
                    "likelyInternalTransfer": internal,
                    "excluded": not is_allowed,
                }
            )
            if is_allowed:
                taken += 1
            else:
                excluded += 1

        print(f"{sheet_name}: в платежи {taken}, помечено исключёнными {excluded}")

    arguments.output.write_text(json.dumps(items, ensure_ascii=False), encoding="utf-8")
    print(f"\nвсего строк: {len(items)}, записано в {arguments.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
