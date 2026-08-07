"""Читает лист СчетаРуб из Buyers.xlsx и складывает строки в JSON для загрузчика.

Раньше этот шаг делался вручную. Теперь это скрипт, потому что перед каждой
перезаливкой файл надо прогнать через проверку, а проверка ест JSON.

Колонки берём по заголовкам, а не по номерам: в файле их 31, порядок со
временем меняется, и привязка к букве колонки означала бы тихую поломку при
любой вставке столбца. Заголовки лежат не в первой строке — над ними две
строки с итогами, поэтому строку заголовков ищем по содержимому.

Ничего не пишет в базу. На выходе — файл JSON.

Использование:
    python read_buyers_xlsx.py Buyers.xlsx deals.json
    python read_buyers_xlsx.py Buyers.xlsx deals.json --period-from 2026-01-01
"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SHEET = "СчетаРуб"

# заголовок в файле → имя поля, которое ждёт загрузчик
COLUMNS = {
    "клиент": "customer",
    "вид документа": "documentType",
    "связанный документ": "documentNumber",
    "дата документа": "documentDate",
    "сумма": "dealAmount",
    "оплачено": "paidAmount",
    "дата оплаты": "paymentDate",
    "остаток по счету": "balance",
    "пояснение": "notes",
    "состояние отгрузки": "shippingStatus",
    "дата поставки": "deliveryDate",
    "наличие закупки": "hasPurchase",
    "менеджер": "manager",
}


def as_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # номер документа приходит числом: 3089.0 в тексте выглядел бы как «3089.0»
        return str(int(value))
    text = str(value).strip()
    return text or None


def as_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def as_number(value: object) -> float | None:
    if value is None or isinstance(value, (datetime, date)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def find_header(sheet) -> tuple[int, dict[int, str]]:
    """Ищем строку заголовков по содержимому: над ней в файле стоят итоги."""
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        titles = {
            str(value).strip().lower(): position
            for position, value in enumerate(row, start=1)
            if value is not None
        }
        if "клиент" in titles and "вид документа" in titles:
            mapping = {
                position: COLUMNS[title]
                for title, position in titles.items()
                if title in COLUMNS
            }
            return index, mapping
    raise SystemExit(
        f"В листе «{SHEET}» не нашлась строка заголовков с колонками «Клиент» и «Вид документа»"
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument(
        "--period-from",
        default="2026-01-01",
        help="брать строки с датой оплаты не раньше этой",
    )
    arguments = parser.parse_args()

    since = date.fromisoformat(arguments.period_from)
    book = load_workbook(arguments.workbook, read_only=True, data_only=True)
    if SHEET not in book.sheetnames:
        raise SystemExit(f"В книге нет листа «{SHEET}»: {book.sheetnames}")
    sheet = book[SHEET]

    header_row, mapping = find_header(sheet)
    missing = set(COLUMNS.values()) - set(mapping.values())
    if missing:
        print(f"в файле не нашлись колонки: {', '.join(sorted(missing))}")

    items: list[dict] = []
    skipped_no_customer = 0
    skipped_by_period = 0

    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index <= header_row:
            continue
        record: dict[str, object] = {"excelRow": index}
        for position, field in mapping.items():
            record[field] = row[position - 1] if position <= len(row) else None

        if not as_text(record.get("customer")):
            skipped_no_customer += 1
            continue

        payment_date = as_date(record.get("paymentDate"))
        if payment_date is None or date.fromisoformat(payment_date) < since:
            skipped_by_period += 1
            continue

        items.append(
            {
                "excelRow": index,
                "customer": as_text(record.get("customer")),
                "documentType": as_text(record.get("documentType")),
                "documentNumber": as_text(record.get("documentNumber")),
                "documentDate": as_date(record.get("documentDate")),
                "dealAmount": as_number(record.get("dealAmount")) or 0,
                "paidAmount": as_number(record.get("paidAmount")) or 0,
                "paymentDate": payment_date,
                "balance": as_number(record.get("balance")) or 0,
                "shippingStatus": as_text(record.get("shippingStatus")),
                "deliveryDate": as_date(record.get("deliveryDate")),
                "hasPurchase": as_text(record.get("hasPurchase")),
                "manager": as_text(record.get("manager")),
                "notes": as_text(record.get("notes")),
            }
        )

    arguments.output.write_text(
        json.dumps(items, ensure_ascii=False), encoding="utf-8"
    )
    print(f"строка заголовков: {header_row}")
    print(f"взято строк: {len(items)}")
    print(f"пропущено без клиента: {skipped_no_customer}, вне периода: {skipped_by_period}")
    print(f"записано в {arguments.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
